from __future__ import annotations

import csv
import hashlib
import json
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_XLSX = Path(r"C:\Users\edwin\Downloads\INFORME_COSTOS_EKA_v6.xlsx")
OUTPUT_DIR = Path(r"D:\CODEX\EKA MINING SAC\exports")
SHEET_NAME = "DETALLE DE REQUERIMIENTOS"


def normalize_header(header: str) -> str:
    mapping = {
        "CLIENTE": "cliente",
        "UNIDAD": "unidad",
        "SOLICITANTE": "solicitante",
        "COTIZACIÓN": "cotizacion_codigo",
        "CENTRO COSTOS": "centro_costos",
        "FECHA RQ": "fecha_rq",
        "N° RQ": "rq_codigo",
        "TIPO": "tipo",
        "TIPO__LIMPIO": "tipo_limpio",
        "TIPO__OBS": "tipo_obs",
        "DESCRIPCIÓN": "descripcion",
        "UND": "und",
        "CANT RQ": "cant_rq",
        "AJUSTE": "ajuste",
        "ATENCION REAL": "atencion_real",
        "CANT. STOCK": "cant_stock",
        "COMPRA": "compra",
        "COSTO UNITARIO \n DÓLAR": "costo_unitario_dolar",
        "COSTO UNITARIO \n SOLES": "costo_unitario_soles",
        "TC": "tc",
        "P.U. SOLES SIN IGV": "pu_soles_sin_igv",
        "COSTO TOTAL PRESUPUESTADO\n[S/]": "costo_total_presupuestado_s",
        "COSTO TOTAL PRESUPUESTADO\n[USD]": "costo_total_presupuestado_usd",
        "FECHA DE COTI": "fecha_coti",
        "ESTADO": "estado",
    }
    return mapping.get(header, header.lower().replace(" ", "_"))


def format_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if " " in text:
        text = text.split(" ")[0]
    return text


def format_scalar(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def first_non_empty(values):
    for value in values:
      if value not in (None, ""):
        return value
    return ""


def main() -> int:
    if not SOURCE_XLSX.exists():
        print(f"No existe el archivo: {SOURCE_XLSX}")
        return 1

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    index = {header: idx for idx, header in enumerate(headers)}

    detail_rows = []
    grouped = OrderedDict()

    for row in rows[1:]:
        raw = {normalize_header(headers[i]): format_scalar(row[i]) for i in range(len(headers))}
        if not raw.get("rq_codigo"):
            continue

        detail_key = hashlib.sha1(
            json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True).encode("utf-8")
        ).hexdigest()

        detail_record = {
            "cotizacion_codigo": raw.get("cotizacion_codigo", ""),
            "centro_costos": str(raw.get("centro_costos", "") or ""),
            "rq_codigo": raw.get("rq_codigo", ""),
            "cliente": raw.get("cliente", ""),
            "unidad": raw.get("unidad", ""),
            "solicitante": raw.get("solicitante", ""),
            "fecha_rq": format_date(raw.get("fecha_rq")),
            "tipo": raw.get("tipo", ""),
            "tipo_limpio": raw.get("tipo_limpio", ""),
            "tipo_obs": raw.get("tipo_obs", ""),
            "descripcion": raw.get("descripcion", ""),
            "und": raw.get("und", ""),
            "cant_rq": raw.get("cant_rq", ""),
            "ajuste": raw.get("ajuste", ""),
            "atencion_real": raw.get("atencion_real", ""),
            "cant_stock": raw.get("cant_stock", ""),
            "compra": raw.get("compra", ""),
            "costo_unitario_dolar": raw.get("costo_unitario_dolar", ""),
            "costo_unitario_soles": raw.get("costo_unitario_soles", ""),
            "tc": raw.get("tc", ""),
            "pu_soles_sin_igv": raw.get("pu_soles_sin_igv", ""),
            "costo_total_presupuestado_s": raw.get("costo_total_presupuestado_s", ""),
            "costo_total_presupuestado_usd": raw.get("costo_total_presupuestado_usd", ""),
            "fecha_coti": format_date(raw.get("fecha_coti")),
            "estado": raw.get("estado", ""),
            "import_batch": SOURCE_XLSX.stem,
            "source_row_hash": detail_key,
            "raw_row": json.dumps(raw, ensure_ascii=False, default=str),
        }
        detail_rows.append(detail_record)

        group_key = (
            detail_record["cotizacion_codigo"],
            detail_record["centro_costos"],
            detail_record["rq_codigo"],
        )
        grouped.setdefault(group_key, []).append(detail_record)

    log_rows = []
    for (cotizacion_codigo, centro_costos, rq_codigo), items in grouped.items():
        log_rows.append(
            {
                "cotizacion_codigo": cotizacion_codigo,
                "centro_costos": centro_costos,
                "rq_codigo": rq_codigo,
                "cliente": first_non_empty(item["cliente"] for item in items),
                "unidad": first_non_empty(item["unidad"] for item in items),
                "solicitante": first_non_empty(item["solicitante"] for item in items),
                "fecha_rq": first_non_empty(item["fecha_rq"] for item in items),
                "tipo_principal": first_non_empty(item["tipo"] for item in items),
                "estado": first_non_empty(item["estado"] for item in items),
                "descripcion_resumen": first_non_empty(item["descripcion"] for item in items),
                "cantidad_items": len(items),
                "origen": "excel",
                "observaciones": "",
                "prioridad": "",
            }
        )

    detail_output = OUTPUT_DIR / "requerimiento_items_import.csv"
    log_output = OUTPUT_DIR / "requerimientos_log_import.csv"

    with detail_output.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(detail_rows[0].keys()))
        writer.writeheader()
        writer.writerows(detail_rows)

    with log_output.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(log_rows[0].keys()))
        writer.writeheader()
        writer.writerows(log_rows)

    print(f"Items exportados: {len(detail_rows)} -> {detail_output}")
    print(f"Requerimientos únicos exportados: {len(log_rows)} -> {log_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
